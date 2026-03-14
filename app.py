import streamlit as st
import pandas as pd
from datetime import datetime
import os
import io
import openpyxl
from openpyxl.drawing.image import Image
from pathlib import Path
import json
import os
from openpyxl.styles import PatternFill

# ---------------- PAGE CONFIG ----------------
st.set_page_config(layout="wide", page_title="SAG ALSAHRA - Technical System")

# ---------------- PATHS ----------------
LOGO_PATH = r"C:\Users\User\Desktop\MyMudMotorApp\SAG_ALSAHRA_LOGO.png"
MWD_HEADER = r"C:\Users\User\Desktop\MyMudMotorApp\MWD HEADER.png"
MOTOR_HEADER = r"C:\Users\User\Desktop\MyMudMotorApp\MUD MOTORHEADER.png"

# Templates
MPI_TEMPLATE = r"C:\Users\User\Desktop\MyMudMotorApp\templates\mwd_lwd_report.xlsx"
LPT_TEMPLATE = r"C:\Users\User\Desktop\MyMudMotorApp\templates\speciality_tool_report.xlsx"
RB_TEMPLATE = r"C:\Users\User\Desktop\MyMudMotorApp\templates\mud motor templates\RB MUD MOTOR.xlsx"
RS_TEMPLATE = r"C:\Users\User\Desktop\MyMudMotorApp\templates\mud motor templates\RS MUD MOTOR.xlsx"

# Output folders
REPORT_OUTPUT = r"C:\Users\User\Desktop\MyMudMotorApp\GENERATED REPORTS"
os.makedirs(REPORT_OUTPUT, exist_ok=True)
PDF_OUTPUT = r"C:\Users\User\Desktop\MyMudMotorApp\SYSTEM REPORTS"
os.makedirs(PDF_OUTPUT, exist_ok=True)
SESSION_SAVE_PATH = Path("session_state.json")

# ---------------- SIMPLE AUTH ----------------
VALID_USER = st.secrets.get("APP_USER", os.environ.get("APP_USER", "sagreport"))
VALID_PASS = st.secrets.get("APP_PASS", os.environ.get("APP_PASS", "inspection"))

if "authed" not in st.session_state:
    st.session_state.authed = False

if not st.session_state.authed:
    st.title("Login")
    user = st.text_input("Username")
    pwd = st.text_input("Password", type="password")
    if st.button("Sign in"):
        if user == VALID_USER and pwd == VALID_PASS:
            st.session_state.authed = True
            st.experimental_rerun()
        else:
            st.error("Invalid credentials")
    st.stop()

# ---------------- UI STYLE ----------------
st.markdown(
    """
<style>
.stApp {background-color:#708090;}
.report-card {background:white;padding:25px;border-radius:10px;color:black;}
[data-testid="stSidebar"] {background:#002b55;}
[data-testid="stSidebar"] * {color:#ffffff;}
</style>
""",
    unsafe_allow_html=True,
)

# ---------------- TABLE STRUCTURES ----------------
TECHNICAL_COLS = [
    "S/N No",
    "Description",
    "Length (MTRS)",
    "Body Cond:","Box Condition","Box_Conn",
    "Box_OD",
    "Box_BB_Dia",
    "Box_BB_Len",
    "Box_CB_Depth",
    "Box_CB_Dia",
    "Box_Bevel",
    "Box_Tong",
    "Box_Thread_Len","Pin Condition","Pin_Conn",
    "Pin_SRG_Dia",
    "Pin_SRG_Len",
    "Pin_Cyl_Dia",
    "Pin_Thread_Len",
    "Pin_Tong",
    "Pin_Bevel",
    "Condition",
]

SPEC_COLS = [
    "S/N No",
    "Description",
    "Length (MTRS)",
    "Body Cond",
    "Box Condition",
    "Box_Conn",
    "Box_OD",
    "Box_BB_Dia",
    "Box_BB_Len",
    "CB_Depth",
    "CB_Dia",
    "Bevel",
    "Box_Tong",
    "Pin_Conn",
    "Pin_OD",
    "Pin_ID",
    "SRG_Dia",
    "SRG_Length",
    "Pin_Cyl_Dia",
    "Pin_Thread_Len",
    "Pin_Tong",
    "Bevel Dia",
    "Condition",
]

# ---------------- HELPERS ----------------

def load_defaults_from_template(template_path, columns, start_row=5):
    wb = openpyxl.load_workbook(template_path, data_only=True)
    sheet = wb.active
    rows = []
    row_idx = start_row
    while True:
        first_col = sheet.cell(row_idx, 1).value
        second_col = sheet.cell(row_idx, 2).value
        if first_col is None and second_col is None:
            break
        row_vals = [sheet.cell(row_idx, col_idx + 1).value for col_idx in range(len(columns))]
        rows.append(row_vals)
        row_idx += 1
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns).fillna("")


def export_excel_to_pdf(xlsx_path, pdf_path):
    try:
        import win32com.client  # type: ignore
    except Exception as exc:
        st.warning(f"PDF export skipped; install pywin32 if you need PDF output. ({exc})")
        return None

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    try:
        workbook = excel.Workbooks.Open(xlsx_path)
        workbook.ExportAsFixedFormat(0, pdf_path)
    finally:
        try:
            workbook.Close(False)
        except Exception:
            pass
        excel.Quit()
    return pdf_path


def add_photo_to_sheet(sheet, image_bytes, anchor_cell, width=300, height=200, label=None):
    img = Image(io.BytesIO(image_bytes))
    img.width = width
    img.height = height
    sheet.add_image(img, anchor_cell)
    if label:
        label_row = max(1, sheet[anchor_cell].row - 1)
        label_col = sheet[anchor_cell].column
        safe_set(sheet, label_row, label_col, label)


def place_motor_photos(photo_sheet, master_df, photo_store):
    if photo_sheet is None:
        return
    col_letters = ["A", "E"]
    start_row = 4
    row_step = 12
    placed = 0
    for idx, row in master_df.iterrows():
        desc = row.get("Description", "")
        sn = row.get("S/N No", idx + 1)
        img_bytes = photo_store.get(desc)
        if not img_bytes:
            continue
        col = col_letters[placed % 2]
        row_num = start_row + (placed // 2) * row_step
        anchor = f"{col}{row_num}"
        label = f"{sn} - {desc}"
        add_photo_to_sheet(photo_sheet, img_bytes, anchor, width=280, height=180, label=label)
        placed += 1


def ensure_serials_as_text():
    for name in ("master_df", "spec_df_ndt", "spec_df_mwd"):
        df = st.session_state.get(name)
        if df is not None and "S/N No" in df.columns:
            st.session_state[name]["S/N No"] = st.session_state[name]["S/N No"].astype(str)


def ensure_columns(name, columns):
    """Guarantee all columns exist and are ordered."""
    df = st.session_state.get(name, pd.DataFrame())
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    st.session_state[name] = df[columns]


def safe_set(sheet, row, col, value):
    """Assign to a cell without breaking merged ranges."""
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            # always write to the top-left cell of the merged range
            sheet.cell(merged.min_row, merged.min_col).value = value
            return
    sheet.cell(row, col).value = value


YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


def apply_condition_color(sheet, coord, value, yellow_codes=None, red_words=None):
    """Color cell based on content rules."""
    text = (str(value) or "").upper().strip()
    if not text:
        return
    yellow_codes = yellow_codes or []
    red_words = red_words or []
    cell = sheet[coord]
    if any(word in text for word in red_words):
        cell.fill = RED_FILL
    elif any(text == code or text.startswith(code) for code in yellow_codes):
        cell.fill = YELLOW_FILL


def load_saved_session():
    if SESSION_SAVE_PATH.exists():
        try:
            data = json.loads(SESSION_SAVE_PATH.read_text())
            st.session_state.master_df = pd.DataFrame(data.get("master_df", []))
            st.session_state.spec_df_ndt = pd.DataFrame(data.get("spec_df_ndt", []))
            st.session_state.spec_df_mwd = pd.DataFrame(data.get("spec_df_mwd", []))
            ensure_columns("master_df", TECHNICAL_COLS)
            ensure_columns("spec_df_ndt", SPEC_COLS)
            ensure_columns("spec_df_mwd", SPEC_COLS)
            ensure_serials_as_text()
        except Exception as exc:
            st.warning(f"Could not load saved session data: {exc}")


def save_session_data():
    payload = {
        "master_df": st.session_state.master_df.to_dict(orient="records"),
        "spec_df_ndt": st.session_state.spec_df_ndt.to_dict(orient="records"),
        "spec_df_mwd": st.session_state.spec_df_mwd.to_dict(orient="records"),
    }
    SESSION_SAVE_PATH.write_text(json.dumps(payload))


# ---------------- MOTOR PART LIST ----------------

def get_motor_parts_list(prefix):
    if prefix == "RB":
        parts = [
            "ROTOR ( B x B )",
            "STATOR ( B x B )",
            "THRUST HOUSING ( B x B )",
            "BEARING ADAPTOR ( B x B )",
            "OFFSET HOUSING ( B x B )",
            "ADAPTOR HOUSING",
            "SPLINE MANDREL",
            "ROTOR ADAPTOR",
            "DRIVE SHAFT",
            "ADJUSTING RING",
            "LOWER SHAFT LOW RESTRICTOR",
            "LOWER HOUSING",
            "STABILIZER SLEEVE",
            "LOCK HOUSING",
            "ROTOR CATCHER",
            "TOP SUB",
            "BEARING MANDREL",
        ]
    elif prefix == "RS":
        parts = [
            "ROTOR ( B x B )",
            "STATOR ( B x B )",
            "BEARING HOUSING ( B x B )",
            "BEARING ADAPTOR ( B x B )",
            "OFFSET HOUSING ( B x B )",
            "STATOR ADAPTOR",
            "SPLINED MANDREL ( P x P )",
            "ROTOR ADAPTOR",
            "DRIVE SHAFT",
            "ADJUSTING RING",
            "CATCHER ROD",
            "STABILIZER SLEEVE",
            "TOP SUB / CATCHER SUB",
            "LOWER ROTATING RADIAL BEARING SLEEVE",
            "LOWER ROTATING RADIAL BEARING RING",
            "UPPER ROTATING RADIAL BEARING SLEEVE",
            "UPPER ROTATING RADIAL BEARING RING",
            "BEARING MANDREL",
        ]
    else:
        return pd.DataFrame(columns=TECHNICAL_COLS)

    spacer_parts_rb = {
        "ROTOR ( B x B )",
        "STATOR ( B x B )",
        "THRUST HOUSING ( B x B )",
        "BEARING ADAPTOR ( B x B )",
        "OFFSET HOUSING ( B x B )",
        "ADAPTOR HOUSING",
        "SPLINE MANDREL",
    }
    spacer_parts_rs = {
        "ROTOR ( B x B )",
        "STATOR ( B x B )",
        "BEARING HOUSING ( B x B )",
        "BEARING ADAPTOR ( B x B )",
        "OFFSET HOUSING ( B x B )",
        "SPLINED MANDREL ( P x P )",
    }
    spacer_targets = spacer_parts_rb if prefix == "RB" else spacer_parts_rs

    rows = []
    for i, p in enumerate(parts, 1):
        row = {col: "" for col in TECHNICAL_COLS}
        row["S/N No"] = i
        row["Description"] = p
        rows.append(row)
        if p in spacer_targets:
            rows.append({col: "" for col in TECHNICAL_COLS})

    return pd.DataFrame(rows)


# ---------------- SESSION STATE ----------------
if "master_df" not in st.session_state:
    st.session_state.master_df = pd.DataFrame(columns=TECHNICAL_COLS)
if "spec_df_ndt" not in st.session_state:
    st.session_state.spec_df_ndt = pd.DataFrame(columns=SPEC_COLS)
if "spec_df_mwd" not in st.session_state:
    st.session_state.spec_df_mwd = pd.DataFrame(columns=SPEC_COLS)
if "uploaded_photos" not in st.session_state:
    st.session_state.uploaded_photos = {}
if "tool_photo_ndt" not in st.session_state:
    st.session_state.tool_photo_ndt = None
if "tool_photo_mwd" not in st.session_state:
    st.session_state.tool_photo_mwd = None

load_saved_session()
ensure_serials_as_text()
ensure_columns("master_df", TECHNICAL_COLS)
ensure_columns("spec_df_ndt", SPEC_COLS)
ensure_columns("spec_df_mwd", SPEC_COLS)

# ---------------- SIDEBAR ----------------
st.sidebar.image("SAG_ALSAHRA_LOGO.png", use_container_width=True)
st.sidebar.title("MAIN MENU")
page = st.sidebar.radio(
    "Go To",
    ["Mud Motor Inspection", "Mud Motor Photo Report", "Specialty Tools (NDT)", "MWD/LWD Tool Report"],
)
job_no = st.sidebar.text_input("JOB NO", "26028")

# ============================================================
# MUD MOTOR INSPECTION
# ============================================================
if page == "Mud Motor Inspection":
    st.image(MOTOR_HEADER, use_container_width=True)
    st.subheader("Inspection Header")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        customer = st.text_input("Customer", "RAY INTERNATIONAL OIL & GAS")
        location = st.text_input("Location", "NIZWA WORKSHOP")
    with c2:
        date_val = st.date_input("Date", datetime.now())
    with c3:
        report_no = st.text_input("Report #", f"SAG-2026-{job_no}-RAY-")
    with c4:
        rig = st.text_input("Rig", "N/A")

    comment_mm = st.text_area("Comments (goes to B51/B52)", "", height=80)

    m1, m2, m3 = st.columns(3)
    with m1:
        motor_id = st.text_input("Mud Motor Serial", "RB962004")
    with m2:
        motor_size = st.selectbox("Size", ["6 3/4\"", "4 3/4\"", "8\"", "9 5/8\""])

    prefix = None
    if motor_id.upper().startswith("RB"):
        prefix = "RB"
    elif motor_id.upper().startswith("RS"):
        prefix = "RS"

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Generate RB Motor Parts"):
            st.session_state.master_df = get_motor_parts_list("RB")
            ensure_serials_as_text()
            save_session_data()
    with col2:
        if st.button("Generate RS Motor Parts"):
            st.session_state.master_df = get_motor_parts_list("RS")
            ensure_serials_as_text()
            save_session_data()

    if st.button("Auto-fill dimensions from template"):
        if prefix:
            template_path = RB_TEMPLATE if prefix == "RB" else RS_TEMPLATE
            st.session_state.master_df = load_defaults_from_template(template_path, TECHNICAL_COLS)
            ensure_serials_as_text()
            save_session_data()
            st.success("Loaded dimensions from template")
        else:
            st.warning("Set a Mud Motor Serial starting with RB or RS to pick the correct template.")

    st.subheader("Technical Inspection")
    f1, f2 = st.columns(2)
    with f1:
        part_search = st.text_input("Search part (S/N or Description)", "")
    with f2:
        motor_filter = st.text_input("Filter by Mud Motor Serial (optional)", motor_id)

    ensure_serials_as_text()
    master_df = st.session_state.master_df
    mask = pd.Series([True] * len(master_df))
    if part_search:
        q = part_search.lower()
        mask &= master_df["S/N No"].astype(str).str.lower().str.contains(q) | master_df[
            "Description"
        ].astype(str).str.lower().str.contains(q)
    filtered_df = master_df[mask]

    edited_data = st.data_editor(
        filtered_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        key="master_editor",
    )
    st.session_state.master_df.loc[edited_data.index] = edited_data
    ensure_serials_as_text()
    save_session_data()

    if st.button("SAVE DATA"):
        save_session_data()
        st.success("Saved")

    if st.button("GENERATE MUD MOTOR REPORT"):
        template = RB_TEMPLATE if prefix == "RB" else RS_TEMPLATE
        wb = openpyxl.load_workbook(template)
        sheet = wb.active
        photo_sheet = wb.worksheets[1] if len(wb.worksheets) > 1 else None

        col_map = {
            "S/N No": 1,              # A25..A48
            "Description": 2,         # B25..B48
            "Length (MTRS)": 3,       # C25..C48
            "Body Cond": 4,           # D25..D48
            "Box_Conn": 5,            # E25..E48
            "Box_OD": 7,              # G25..G48
            "Box_BB_Dia": 8,          # H25..H48
            "Box_BB_Len": 9,          # I25..I48
            "Box_CB_Depth": 10,       # J25..J48
            "Box_CB_Dia": 11,         # K25..K48
            "Box_Bevel": 12,          # L25..L48
            "Box_Tong": 13,           # M25..M48
            "Box_Thread_Len": 14,     # N25..N48
            "Pin_Conn": 15,           # O25..O48
            "Pin_SRG_Dia": 16,        # P25..P48
            "Pin_SRG_Len": 17,        # Q25..Q48
            "Pin_Cyl_Dia": 18,        # R25..R48
            "Pin_Thread_Len": 21,     # U25..U48
            "Pin_Tong": 22,           # V25..V48
            "Pin_Bevel": 23,          # W25..W48
            "Pin_OD": 24,             # X25..X48
            "Pin_ID": 25,             # Y25..Y48
            "Condition": 26,          # Z25..Z48
        }
        start_row = 25
        for idx, row in enumerate(st.session_state.master_df.itertuples(index=False), start_row):
            for field, col_idx in col_map.items():
                if field in st.session_state.master_df.columns:
                    val = row[st.session_state.master_df.columns.get_loc(field)]
                    safe_set(sheet, idx, col_idx, val)

        safe_set(sheet, 6, 2, customer)           # B6 customer
        safe_set(sheet, 6, 12, location)          # L6 location
        safe_set(sheet, 6, 18, job_no)            # R6 job number
        safe_set(sheet, 5, 24, report_no)         # X5 report number
        safe_set(sheet, 6, 24, str(date_val))     # X6 date
        safe_set(sheet, 8, 3, "✔")                # C8 tick mark
        safe_set(sheet, 20, 3, motor_id)          # C20 mud motor serial
        safe_set(sheet, 20, 19, motor_size)       # S20 tool size
        safe_set(sheet, 61, 13, str(date_val))    # M61 duplicate date
        if comment_mm:
            lines = comment_mm.splitlines()
            sheet["B51"] = lines[0]
            if len(lines) > 1:
                sheet["B52"] = lines[1]

        # Row 56 summary counts
        rows_limited = st.session_state.master_df.head(18).reset_index(drop=True)
        expected = 18

        def norm(val):
            return str(val).strip().upper() if val is not None else ""

        def is_blank_row(r):
            return all(norm(r.get(col, "")) == "" for col in ["Description"])

        blank_rows = sum(is_blank_row(r) for r in rows_limited.to_dict(orient="records"))
        filled_rows = expected - blank_rows

        not_ok_rows = 0
        repairable = 0
        scrap_dbr = 0
        repairable_codes = {"GT", "TD", "DT", "WT", "PT", "PSRG", "CS", "SD", "CONN ID PIT", "WO", "SB", "BB", "TS SHORT", "DS"}

        for r in rows_limited.to_dict(orient="records"):
            body = norm(r.get("Body Cond", ""))
            boxc = norm(r.get("Box Condition", ""))
            pinc = norm(r.get("Condition", ""))

            if not (body == "OK" and boxc == "OK" and pinc == "OK"):
                not_ok_rows += 1

            if ("SCRAP" in boxc or "DBR" in boxc or "SCRAP" in pinc or "DBR" in pinc):
                scrap_dbr += 1
            else:
                if boxc in repairable_codes or pinc in repairable_codes:
                    repairable += 1

        safe_set(sheet, 56, 2, expected - not_ok_rows)  # B56 = 18 - not_ok
        safe_set(sheet, 56, 8, repairable)              # H56 repairable count
        safe_set(sheet, 56, 15, scrap_dbr)              # O56 scrap/DBR count
        safe_set(sheet, 56, 24, filled_rows)            # X56 filled count

        place_motor_photos(photo_sheet, st.session_state.master_df, st.session_state.uploaded_photos)

        filename = f"MUD_MOTOR_{motor_id}_{date_val}.xlsx"
        save_path = os.path.join(REPORT_OUTPUT, filename)
        wb.save(save_path)
        save_session_data()

        st.success("Report Generated Successfully")
        st.info(f"Saved to: {save_path}")

        pdf_path = os.path.join(PDF_OUTPUT, filename.replace(".xlsx", ".pdf"))
        pdf_result = export_excel_to_pdf(save_path, pdf_path)

        with open(save_path, "rb") as f:
            st.download_button("Download Report", f, file_name=filename)
        if pdf_result:
            with open(pdf_result, "rb") as f:
                st.download_button("Download PDF", f, file_name=os.path.basename(pdf_result))

# ============================================================
# PHOTO REPORT
# ============================================================
elif page == "Mud Motor Photo Report":
    st.image(MOTOR_HEADER, use_container_width=True)
    st.subheader("Component Photos")

    # capture basic info for headers
    photo_motor_id = st.text_input("Motor Serial (for photo report)", st.session_state.get("photo_motor_id", ""))
    st.session_state.photo_motor_id = photo_motor_id
    photo_motor_size = st.selectbox(
        "Motor Size (for photo report)",
        ['6 3/4"', '4 3/4"', '8"', '9 5/8"'],
        index=0 if st.session_state.get("photo_motor_size") is None else
        max(0, ['6 3/4"', '4 3/4"', '8"', '9 5/8"'].index(st.session_state.get("photo_motor_size", '6 3/4"')))
    )
    st.session_state.photo_motor_size = photo_motor_size
    comment_line1 = st.text_input("Photo report comment line 1 (B54)", st.session_state.get("photo_comment1", ""))
    comment_line2 = st.text_input("Photo report comment line 2 (B55)", st.session_state.get("photo_comment2", ""))
    st.session_state.photo_comment1 = comment_line1
    st.session_state.photo_comment2 = comment_line2

    if not st.session_state.master_df.empty:
        parts = st.session_state.master_df
        cols = st.columns(4)
        for i, row in parts.iterrows():
            with cols[i % 4]:
                part = str(row.get("Description", "")).strip()
                if not part:
                    continue
                file = st.file_uploader(part, key=f"photo_{i}_{part}")
                if file:
                    img_bytes = file.read()
                    st.image(img_bytes, width=200)
                    st.session_state.uploaded_photos[part] = img_bytes
    else:
        st.info("Add mud motor parts first in the inspection tab.")

    # Header fields for photo sheet (second sheet)
    if st.button("Update Photo Sheet Headers"):
        try:
            template = RB_TEMPLATE if prefix == "RB" else RS_TEMPLATE
            wb = openpyxl.load_workbook(template)
            photo_sheet = wb.worksheets[1] if len(wb.worksheets) > 1 else None
            if photo_sheet:
                safe_set(photo_sheet, 6, 2, customer)        # B6 customer
                safe_set(photo_sheet, 6, 12, location)       # L6 location
                safe_set(photo_sheet, 6, 18, job_no)         # R6 job number
                safe_set(photo_sheet, 5, 24, report_no)      # X5 report number
                safe_set(photo_sheet, 6, 24, str(date_val))  # X6 date
                safe_set(photo_sheet, 56, 13, str(date_val)) # M56 date
                safe_set(photo_sheet, 8, 3, "✔")             # C8 tick mark
                safe_set(photo_sheet, 12, 3, photo_motor_id) # C12 motor number
                safe_set(photo_sheet, 12, 19, photo_motor_size) # S12 motor size
                safe_set(photo_sheet, 54, 2, comment_line1)  # B54
                safe_set(photo_sheet, 55, 2, comment_line2)  # B55
                wb.save(template)
                st.success("Photo sheet headers updated for future photo reports.")
            else:
                st.warning("Template has no second sheet to update photo headers.")
        except Exception as exc:
            st.error(f"Could not update photo sheet headers: {exc}")
# ============================================================
# SPECIALTY TOOLS (NDT)
# ============================================================
elif page == "Specialty Tools (NDT)":
    st.image(MWD_HEADER, use_container_width=True)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        tool_type = "LPT"
        customer = st.text_input("Customer", "RAY INTERNATIONAL OIL & GAS")
        rig = st.text_input("Rig", "N/A")
    with c2:
        date_val = st.date_input("Date", datetime.now())
        tool_id = st.text_input("Tool ID", "...")
    with c3:
        report_no = st.text_input("Report", f"SAG-2026-{job_no}-RAY-")
    with c4:
        job_loc = st.text_input("Location", "N/A")

    comment_ndt = st.text_area("Comments (goes to C37)", "", height=80)

    photo_file = st.file_uploader("Upload tool photo (single)", type=["png", "jpg", "jpeg"], key="ndt_photo")
    if photo_file:
        st.session_state.tool_photo_ndt = photo_file.read()
        st.image(st.session_state.tool_photo_ndt, width=220)

    f_ndt1, f_ndt2 = st.columns(2)
    with f_ndt1:
        ndt_search_sn = st.text_input("Search by S/N (NDT)", "")
    with f_ndt2:
        ndt_search_desc = st.text_input("Search by Description (NDT)", "")

    if st.button("Load template values", key="ndt_load_defaults"):
        st.session_state.spec_df_ndt = load_defaults_from_template(LPT_TEMPLATE, SPEC_COLS)
        save_session_data()

    ndt_df = st.session_state.spec_df_ndt
    ndt_mask = pd.Series([True] * len(ndt_df))
    if ndt_search_sn:
        ndt_mask &= ndt_df["S/N No"].astype(str).str.contains(ndt_search_sn, case=False, na=False)
    if ndt_search_desc:
        ndt_mask &= ndt_df["Description"].astype(str).str.contains(ndt_search_desc, case=False, na=False)
    ndt_filtered = ndt_df[ndt_mask]

    edited_spec = st.data_editor(
        ndt_filtered,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        key="ndt_editor",
    )
    # persist edits immediately
    st.session_state.spec_df_ndt.loc[edited_spec.index] = edited_spec
    ensure_serials_as_text()
    save_session_data()

    if st.button("SAVE NDT DATA", key="save_ndt"):
        save_session_data()
        st.success("Saved")

    if st.button("GENERATE NDT REPORT"):
        wb = openpyxl.load_workbook(LPT_TEMPLATE)
        sheet = wb.active

        start_row = 24  # template data row starts at row 24
        col_map = {
            "S/N No": 2,            # B24
            "Description": 3,       # C24
            "Length (MTRS)": 4,     # D24
            "Body Cond": 5,         # E24
            "Box Condition": 15,    # O24
            "Box_Conn": 6,          # F24
            "Box_OD": 7,            # G24
            "Box_BB_Dia": 8,        # H24
            "Box_BB_Len": 9,        # I24
            "CB_Depth": 10,         # J24
            "CB_Dia": 11,           # K24
            "Bevel": 12,            # L24
            "Box_Tong": 13,         # M24
            "Pin_Conn": 16,         # P24
            "SRG_Dia": 17,          # Q24
            "SRG_Length": 18,       # R24
            "Pin_Cyl_Dia": 19,      # S24 (nose/cylinder)
            "Pin_Thread_Len": 22,   # V24
            "Pin_Tong": 23,         # W24
            "Bevel Dia": 24,        # X24
            "Pin_OD": 25,           # Y24
            "Pin_ID": 26,           # Z24
            "Condition": 27,        # AA24
        }

        for idx, row in enumerate(st.session_state.spec_df_ndt.itertuples(index=False), start_row):
            for col_name, col_idx in col_map.items():
                if col_name in st.session_state.spec_df_ndt.columns:
                    val = row[st.session_state.spec_df_ndt.columns.get_loc(col_name)]
                    safe_set(sheet, idx, col_idx, val)

        safe_set(sheet, 6, 3, customer)     # C6
        safe_set(sheet, 6, 13, job_loc)    # M6
        safe_set(sheet, 6, 21, rig)         # U6
        safe_set(sheet, 6, 26, str(date_val))  # Z6
        safe_set(sheet, 8, 21, job_no)      # U8
        safe_set(sheet, 8, 26, report_no)   # Z8
        safe_set(sheet, 8, 4, "✔")          # D8 tick
        if comment_ndt:
            safe_set(sheet, 37, 3, comment_ndt)

        if st.session_state.tool_photo_ndt:
            add_photo_to_sheet(sheet, st.session_state.tool_photo_ndt, "K27", width=758, height=230, label=None)

        # Derived flags and footer/header extras
        first_row = st.session_state.spec_df_ndt.iloc[0] if not st.session_state.spec_df_ndt.empty else None
        box_cond_val = str(first_row.get("Box Condition", "")).strip().upper() if first_row is not None else ""
        pin_cond_val = str(first_row.get("Condition", "")).strip().upper() if first_row is not None else ""
        body_cond_val = str(first_row.get("Body Cond", "")).strip().upper() if first_row is not None else ""
        c39_val = 1 if (box_cond_val == "OK" and pin_cond_val == "OK" and body_cond_val == "OK") else 0
        l39_val = 1 if (box_cond_val != "OK" or pin_cond_val != "OK") else 0
        safe_set(sheet, 39, 3, c39_val)   # C39
        safe_set(sheet, 39, 12, l39_val)  # L39
        safe_set(sheet, 41, 3, sheet.cell(24, 15).value)   # C41 = O24
        safe_set(sheet, 41, 12, sheet.cell(24, 5).value)   # L41 = E24
        safe_set(sheet, 41, 27, sheet.cell(24, 27).value)   # AA41 = AA24

        # coloring rules
        apply_condition_color(sheet, "E24", sheet.cell(24, 5).value, red_words=["CRACK", "CRACKED", "BENT", "DBR"])
        apply_condition_color(sheet, "O24", sheet.cell(24, 15).value,
                              yellow_codes=["DT", "TD", "WT", "GT", "SD", "DS", "CS", "PT", "BB", "SB", "WO"],
                              red_words=["CRACK", "DBR"])
        apply_condition_color(sheet, "AA24", sheet.cell(24, 27).value,
                              yellow_codes=["DT", "TD", "WT", "GT", "SD", "DS", "CS", "PT", "WO"],
                              red_words=["CRACK", "DBR"])
        # mirror colors
        def clone_fill(src):
            return PatternFill(start_color=src.start_color, end_color=src.end_color, fill_type=src.fill_type)
        sheet["C41"].fill = clone_fill(sheet["O24"].fill)
        sheet["L41"].fill = clone_fill(sheet["E24"].fill)
        sheet["AA41"].fill = clone_fill(sheet["AA24"].fill)
        safe_set(sheet, 43, 10, date_val) # J43
        safe_set(sheet, 43, 2, "JIJU JAYAPRASAD")  # B43

        filename = f"LPT_{tool_id}_{date_val}.xlsx"
        save_path = os.path.join(REPORT_OUTPUT, filename)
        wb.save(save_path)
        save_session_data()

        with open(save_path, "rb") as f:
            st.download_button("Download Tool Report", f, file_name=filename)

        pdf_path = os.path.join(PDF_OUTPUT, filename.replace(".xlsx", ".pdf"))
        pdf_result = export_excel_to_pdf(save_path, pdf_path)
        if pdf_result:
            with open(pdf_result, "rb") as f:
                st.download_button("Download PDF", f, file_name=os.path.basename(pdf_result))
# ============================================================
# MWD / LWD TOOL REPORT
# ============================================================
elif page == "MWD/LWD Tool Report":
    st.image(MWD_HEADER, use_container_width=True)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        tool_type = "MPI"
        customer = st.text_input("Customer", "RAY INTERNATIONAL OIL & GAS")
        rig = st.text_input("Rig", "N/A")
    with c2:
        date_val = st.date_input("Date", datetime.now())
        tool_id = st.text_input("Tool ID", "...")
    with c3:
        report_no = st.text_input("Report", f"SAG-2026-{job_no}-RAY-")
    with c4:
        job_loc = st.text_input("Location", "NIZWA WORKSHOP")

    comment_mwd = st.text_area("Comments (goes to C37)", "", height=80)

    photo_file_mwd = st.file_uploader("Upload tool photo (single)", type=["png", "jpg", "jpeg"], key="mwd_photo")
    if photo_file_mwd:
        st.session_state.tool_photo_mwd = photo_file_mwd.read()
        st.image(st.session_state.tool_photo_mwd, width=220)

    f_mwd1, f_mwd2 = st.columns(2)
    with f_mwd1:
        mwd_search_sn = st.text_input("Search by S/N (MWD/LWD)", "")
    with f_mwd2:
        mwd_search_desc = st.text_input("Search by Description (MWD/LWD)", "")

    if st.button("Load template values ", key="mwd_load_defaults"):
        st.session_state.spec_df_mwd = load_defaults_from_template(MPI_TEMPLATE, SPEC_COLS)
        save_session_data()

    mwd_df = st.session_state.spec_df_mwd
    if mwd_df.empty:
        mwd_df = pd.DataFrame([[""] * len(SPEC_COLS)], columns=SPEC_COLS)
        st.session_state.spec_df_mwd = mwd_df
    mwd_mask = pd.Series([True] * len(mwd_df))
    if mwd_search_sn:
        mwd_mask &= mwd_df["S/N No"].astype(str).str.contains(mwd_search_sn, case=False, na=False)
    if mwd_search_desc:
        mwd_mask &= mwd_df["Description"].astype(str).str.contains(mwd_search_desc, case=False, na=False)
    mwd_filtered = mwd_df[mwd_mask]

    edited_spec = st.data_editor(
        mwd_filtered,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        key="mwd_editor",
    )
    # persist edits immediately
    st.session_state.spec_df_mwd.loc[edited_spec.index] = edited_spec
    ensure_serials_as_text()
    save_session_data()

    if st.button("SAVE MWD DATA", key="save_mwd"):
        save_session_data()
        st.success("Saved")

    if st.button("GENERATE MWD REPORT"):
        wb = openpyxl.load_workbook(MPI_TEMPLATE)
        sheet = wb.active

        start_row = 24  # template data row starts at row 24
        col_map = {
            "S/N No": 2,            # B24
            "Description": 3,       # C24
            "Length (MTRS)": 4,     # D24
            "Body Cond": 5,         # E24
            "Box Condition": 15,    # O24
            "Box_Conn": 6,          # F24
            "Box_OD": 7,            # G24
            "Box_BB_Dia": 8,        # H24
            "Box_BB_Len": 9,        # I24
            "CB_Depth": 10,         # J24
            "CB_Dia": 11,           # K24
            "Bevel": 12,            # L24
            "Box_Tong": 13,         # M24
            "Pin_Conn": 16,         # P24
            "SRG_Dia": 17,          # Q24
            "SRG_Length": 18,       # R24
            "Pin_Cyl_Dia": 19,      # S24 (nose/cylinder)
            "Pin_Thread_Len": 22,   # V24
            "Pin_Tong": 23,         # W24
            "Bevel Dia": 24,        # X24
            "Pin_OD": 25,           # Y24
            "Pin_ID": 26,           # Z24
            "Condition": 27,        # AA24
        }

        for idx, row in enumerate(st.session_state.spec_df_mwd.itertuples(index=False), start_row):
            for col_name, col_idx in col_map.items():
                if col_name in st.session_state.spec_df_mwd.columns:
                    val = row[st.session_state.spec_df_mwd.columns.get_loc(col_name)]
                    safe_set(sheet, idx, col_idx, val)

        safe_set(sheet, 6, 3, customer)     # C6
        safe_set(sheet, 6, 13, location)    # M6
        safe_set(sheet, 6, 21, rig)         # U6
        safe_set(sheet, 6, 26, str(date_val))  # Z6
        safe_set(sheet, 8, 18, job_no)      # R8
        safe_set(sheet, 8, 26, report_no)   # Z8
        safe_set(sheet, 8, 4, "✔")          # D8 tick mark
        if comment_mwd:
            safe_set(sheet, 37, 3, comment_mwd)

        if st.session_state.tool_photo_mwd:
            add_photo_to_sheet(sheet, st.session_state.tool_photo_mwd, "K27", width=758, height=230, label=None)

        first_row = st.session_state.spec_df_mwd.iloc[0] if not st.session_state.spec_df_mwd.empty else None
        box_cond_val = str(first_row.get("Box Condition", "")).strip().upper() if first_row is not None else ""
        pin_cond_val = str(first_row.get("Condition", "")).strip().upper() if first_row is not None else ""
        body_cond_val = str(first_row.get("Body Cond", "")).strip().upper() if first_row is not None else ""
        c39_val = 1 if (box_cond_val == "OK" and pin_cond_val == "OK" and body_cond_val == "OK") else 0
        l39_val = 1 if (box_cond_val != "OK" or pin_cond_val != "OK") else 0
        safe_set(sheet, 39, 3, c39_val)   # C39
        safe_set(sheet, 39, 12, l39_val)  # L39
        safe_set(sheet, 41, 3, sheet.cell(24, 15).value)   # C41 = O24
        safe_set(sheet, 41, 12, sheet.cell(24, 5).value)   # L41 = E24
        safe_set(sheet, 41, 27, sheet.cell(24, 27).value)   # AA41 = AA24

        apply_condition_color(sheet, "E24", sheet.cell(24, 5).value, red_words=["CRACK", "CRACKED", "BENT", "DBR"])
        apply_condition_color(sheet, "O24", sheet.cell(24, 15).value,
                              yellow_codes=["DT", "TD", "WT", "GT", "SD", "DS", "CS", "PT", "BB", "SB", "WO"],
                              red_words=["CRACK", "DBR"])
        apply_condition_color(sheet, "AA24", sheet.cell(24, 27).value,
                              yellow_codes=["DT", "TD", "WT", "GT", "SD", "DS", "CS", "PT", "WO"],
                              red_words=["CRACK", "DBR"])
        def clone_fill(src):
            return PatternFill(start_color=src.start_color, end_color=src.end_color, fill_type=src.fill_type)
        sheet["C41"].fill = clone_fill(sheet["O24"].fill)
        sheet["L41"].fill = clone_fill(sheet["E24"].fill)
        sheet["AA41"].fill = clone_fill(sheet["AA24"].fill)
        safe_set(sheet, 43, 10, date_val) # J43
        safe_set(sheet, 43, 2, "JIJU JAYAPRASAD")  # B43

        filename = f"MPI_{tool_id}_{date_val}.xlsx"
        save_path = os.path.join(REPORT_OUTPUT, filename)
        wb.save(save_path)
        save_session_data()

        with open(save_path, "rb") as f:
            st.download_button("Download Tool Report", f, file_name=filename)

        pdf_path = os.path.join(PDF_OUTPUT, filename.replace(".xlsx", ".pdf"))
        pdf_result = export_excel_to_pdf(save_path, pdf_path)
        if pdf_result:
            with open(pdf_result, "rb") as f:
                st.download_button("Download PDF", f, file_name=os.path.basename(pdf_result))



